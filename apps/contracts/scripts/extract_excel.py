#!/usr/bin/env python3
"""Extract the FMSS contract workbook into data/seed.json.

Node has no built-in xlsx reader, so this Python script (openpyxl) is the
authoritative parser of the messy sheet layout. The JSON it emits is the
contract between this script and server/db.js, which loads it on seed.

Re-run any time the source workbook changes:
    python scripts/extract_excel.py
"""
import json
import re
import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Sample_Data" / "365 Sport Contract 2025.xlsx"
OUT = ROOT / "data" / "seed.json"

# Shared rate card (from Sat sheet rows 40-45, confirmed against Mon/Thu charges).
RATES = {
    "contracted_10": 30, "contracted_12": 27,
    "captain_10": 25, "captain_12": 20,
    "noncontract": 35,
}

# Known WhatsApp name variants -> canonical ledger name. The app can add more.
ALIASES = {
    "Tushar": ["Tush"],
    "Jeetu": ["Jithu"],
    "Saheer": ["Sahir"],
    "Khaled": ["Khalid"],
    "Rojy": ["Roji"],
    "Nimesh": ["Nimeesh"],
    "AWS + Ali": ["Aws", "Ali", "Aws Ali"],
}

# Per-sheet layout. Columns are 1-based (openpyxl). gw_cols is inclusive range.
SHEETS = {
    "sat": {
        "tab": "Sat",
        "name": "Saturday Morning Madness",
        "venue": "365 / Danube",
        "cost_per_gw": 292.5,
        "gw_cols": (12, 35),          # L .. AI
        "rows": {"contract_number": 2, "date": 3, "cost_per_gw": 4,
                 "teams": 5, "captains": 6, "score": 7, "num_players": 8},
        "players": (10, 30),          # ledger rows
        "cols": {"idx": 1, "name": 2, "old": 3, "contrib": 4, "present": 5,
                 "status": 7, "played": 8, "deducted": 9},
    },
    "monthu": {
        "tab": "Mon Thu",
        "name": "Monday / Thursday (O365)",
        "venue": "O365",
        "cost_per_gw": 275,
        "gw_cols": (13, 34),          # M .. AH
        "rows": {"contract_number": 2, "date": 3, "cost_per_gw": 5,
                 "teams": 6, "captains": 7, "score": 8, "num_players": 9},
        "players": (11, 37),
        "cols": {"idx": 1, "name": 2, "old": 3, "contrib": 4, "present": 5,
                 "status": 7, "capt_subsidy": 8, "deducted": 9, "played": 10},
    },
}


def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def num(v):
    """Coerce to float if numeric, else None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
    return None


def slug(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    players = {}      # canonical name (lower) -> {id, name, aliases}
    alias_lookup = {}
    for canon, variants in ALIASES.items():
        for v in variants:
            alias_lookup.setdefault(canon, []).append(v)

    def register_player(name):
        name = name.strip()
        key = name.lower()
        if key not in players:
            players[key] = {
                "id": slug(name),
                "name": name,
                "aliases": alias_lookup.get(name, []),
            }
        return players[key]["id"]

    contracts = []
    ledgers = []
    gameweeks = []

    for cid, cfg in SHEETS.items():
        ws = wb[cfg["tab"]]
        R = cfg["rows"]
        C = cfg["cols"]
        contracts.append({
            "id": cid, "name": cfg["name"], "venue": cfg["venue"],
            "cost_per_gw": cfg["cost_per_gw"], "rates": RATES,
            "sort": 0 if cid == "sat" else 1,
        })

        # --- player ledger rows -> opening balances + row->player map ---
        row_player = {}     # ledger row -> player_id
        p0, p1 = cfg["players"]
        for r in range(p0, p1 + 1):
            name = clean(ws.cell(r, C["name"]).value)
            if not name or name.lower().startswith("other"):
                continue
            pid = register_player(name)
            row_player[r] = pid
            present = num(ws.cell(r, C["present"]).value)
            ledgers.append({
                "player_id": pid, "contract_id": cid,
                "opening_balance": round(present, 2) if present is not None else 0.0,
                "status": clean(ws.cell(r, C["status"]).value) or "",
                "played_count": int(num(ws.cell(r, C["played"]).value) or 0),
            })

        # --- gameweeks across the date columns ---
        gc0, gc1 = cfg["gw_cols"]
        gw_no = 0
        for c in range(gc0, gc1 + 1):
            date = clean(ws.cell(R["date"], c).value)
            teams = clean(ws.cell(R["teams"], c).value)
            score = clean(ws.cell(R["score"], c).value)
            num_pl = num(ws.cell(R["num_players"], c).value)
            # Skip empty / "No Game" columns.
            if not date and not teams:
                continue
            if teams and teams.strip().lower() in ("no game", "no games"):
                continue
            gw_no += 1
            cpg = num(ws.cell(R["cost_per_gw"], c).value)
            charges = []
            for r, pid in row_player.items():
                amt = num(ws.cell(r, c).value)
                if amt is not None and amt != 0:
                    charges.append({"player_id": pid, "amount": round(amt, 2)})
            gameweeks.append({
                "id": f"{cid}_gw{gw_no}",
                "contract_id": cid,
                "gw_number": gw_no,
                "contract_number": int(num(ws.cell(R["contract_number"], c).value) or 0),
                "date": date,
                "cost_per_gw": cpg if cpg is not None else cfg["cost_per_gw"],
                "num_players": int(num_pl) if num_pl is not None else len(charges),
                "teams_raw": teams or "",
                "captains_raw": clean(ws.cell(R["captains"], c).value) or "",
                "score": score or "",
                "comments": "",
                "historical": 1,
                "charges": charges,
            })

    # --- contributions (Contri sheet: left A:D 2024-25, right G:J 2026) ---
    contributions = []
    ws = wb["Contri"]
    blocks = [(1, 2, 3, 4), (7, 8, 9, 10)]   # (name, amount, date, comments) cols
    for (cn, ca, cd, cc) in blocks:
        for r in range(2, ws.max_row + 1):
            name = clean(ws.cell(r, cn).value)
            amount = num(ws.cell(r, ca).value)
            if name is None and amount is None:
                continue
            # match to a known player by canonical name (loose)
            pid = None
            if name:
                key = name.strip().lower()
                if key in players:
                    pid = players[key]["id"]
            contributions.append({
                "player_id": pid,
                "name_raw": name or "",
                "amount": round(amount, 2) if amount is not None else 0.0,
                "date": clean(ws.cell(r, cd).value),
                "comments": clean(ws.cell(r, cc).value) or "",
                "historical": 1,
            })

    # --- kitty (Kitty Sheet right block I:N is the current ledger) ---
    kitty = []
    ws = wb["Kitty Sheet"]
    # income items: cols I (label), J (amount)
    for r in range(2, ws.max_row + 1):
        label = clean(ws.cell(r, 9).value)
        amount = num(ws.cell(r, 10).value)
        if label and amount is not None and not label.lower().startswith(
                ("expense", "total", "present balance")):
            kitty.append({"kind": "income", "label": label,
                          "amount": round(amount, 2), "date": None,
                          "scope": "", "historical": 1})
    # expense items: cols L (label), M (amount), N (date/year)
    for r in range(2, ws.max_row + 1):
        label = clean(ws.cell(r, 12).value)
        amount = num(ws.cell(r, 13).value)
        if label and amount is not None and not label.lower().startswith("expense"):
            kitty.append({"kind": "expense", "label": label,
                          "amount": round(amount, 2),
                          "date": clean(ws.cell(r, 14).value),
                          "scope": "", "historical": 1})

    seed = {
        "meta": {
            "import_date": datetime.date.today().isoformat(),
            "source": SRC.name,
            "kitty_opening_balance": 2894.25,   # Kitty Sheet J19 "Present Balance"
        },
        "contracts": contracts,
        "players": [players[k] for k in players],
        "ledgers": ledgers,
        "gameweeks": gameweeks,
        "contributions": contributions,
        "kitty": kitty,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(seed, indent=2, ensure_ascii=False), encoding="utf-8")

    # summary
    print(f"Wrote {OUT}")
    print(f"  players       : {len(seed['players'])}")
    print(f"  ledgers       : {len(seed['ledgers'])}")
    print(f"  gameweeks     : {len(seed['gameweeks'])} "
          f"(sat={sum(1 for g in gameweeks if g['contract_id']=='sat')}, "
          f"monthu={sum(1 for g in gameweeks if g['contract_id']=='monthu')})")
    print(f"  charges       : {sum(len(g['charges']) for g in gameweeks)}")
    print(f"  contributions : {len(seed['contributions'])} "
          f"(total {round(sum(c['amount'] for c in contributions),1)} AED)")
    print(f"  kitty entries : {len(seed['kitty'])}")


if __name__ == "__main__":
    main()
